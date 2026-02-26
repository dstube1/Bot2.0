
from abc import ABC, abstractmethod
import pyautogui
import os
from bot.base import BotState, PlayerInput, InventoryManager, RestartTask, debug, info, warn, error
import time
import json


def load_gacha_boxes_from_config() -> list:
    """Load all gacha boxes from config/gacha_sets.json dynamically.

    Returns a list of dicts like:
      { 'teleporter': <tp_box>, 'pego_view_direction': <view>,
        'gacha1_view_direction': <view>, 'gacha2_view_direction': <view> }
    Missing views are omitted.
    """
    cfg_path = os.path.join(os.path.dirname(__file__), "..", "config", "gacha_sets.json")
    try:
        with open(cfg_path, "r") as f:
            data = json.load(f)
        sets = data.get("gacha_sets", [])
    except Exception as e:
        error(f"load_gacha_boxes_from_config: failed to load: {e}")
        return []
    boxes = []
    for box in sets:
        tp_box = box.get("tp_box")
        # Prefer explicit per-box pego keys, else generic key
        pego_view = (
            box.get("pego_view_direction")
            or box.get("pego1_view_direction")
            or box.get("pego2_view_direction")
            or box.get("pego3_view_direction")
            or box.get("pego4_view_direction")
            or box.get("pego5_view_direction")
        )
        entry = {"teleporter": tp_box}
        if pego_view is not None:
            entry["pego_view_direction"] = pego_view
        # Include optional gacha views for downstream tasks that may use them
        g1_view = box.get("gacha1_view_direction") or box.get("gacha1", {}).get("view_direction")
        g2_view = box.get("gacha2_view_direction") or box.get("gacha2", {}).get("view_direction")
        if g1_view is not None:
            entry["gacha1_view_direction"] = g1_view
        if g2_view is not None:
            entry["gacha2_view_direction"] = g2_view
        boxes.append(entry)
    return boxes


class BaseTask(ABC):
    """Abstract base class for all atomic bot tasks."""
    @abstractmethod
    def run(self):
        pass


class MovePlayerTask(BaseTask):
    def __init__(self, bot_state: BotState, player_input: PlayerInput, target_position):
        self.bot_state: BotState = bot_state
        self.player_input: PlayerInput = player_input
        self.target_position = target_position

    def run(self):
        self.player_input.log_task(self.bot_state, f"MovePlayerTask -> {self.target_position}", task_obj=self)
        self.player_input.teleport_to(self.target_position, self.bot_state)


class GetTrapsTask(BaseTask):
    """Collect traps from specified crop plots into player inventory.

    Expects each plot dict to contain:
    - 'position': teleport target for the plot
    - 'view_direction': tuple used for `look_at`

    Usage:
        GetTrapsTask(bot_state, player_input, inventory_manager, crop_plots, indices=range(0,6)).run()
    """
    def __init__(self, bot_state: BotState, player_input: PlayerInput, inventory_manager: InventoryManager, crop_plots, indices=None):
        self.bot_state: BotState = bot_state
        self.player_input: PlayerInput = player_input
        self.inventory_manager: InventoryManager = inventory_manager
        self.crop_plots = crop_plots
        # indices: iterable of plot indices to gather from; None means all
        self.indices = list(indices) if indices is not None else list(range(len(crop_plots)))

    def run(self):
        self.player_input.log_task(self.bot_state, f"GetTrapsTask indices={list(self.indices)}", task_obj=self)
        # Caller teleports to plots area beforehand.
        for i in self.indices:
            plot = self.crop_plots[i]
            debug(f"Collecting trap from plot {i} with view {plot.get('view_direction')}")
         
            # Align crouch state with plot requirement
            required_crouch = bool(plot.get('crouch', False))
            if required_crouch != bool(self.bot_state.is_crouching):
                self.player_input.crouch(self.bot_state)
            self.player_input.look_at(plot['view_direction'], self.bot_state)
            self.inventory_manager.open_inv()
            # Assumes 'trap' is the item to pull from each plot
            self.inventory_manager.take_item("trap")
            self.inventory_manager.close_inv()
            time.sleep(0.2)  # small delay between plots
        # After collecting all traps, stand up if crouched to restore neutral state
        if getattr(self.bot_state, 'is_crouching', False):
            self.player_input.crouch(self.bot_state)


class FeedGachaTask(BaseTask):
    """Teleport to a gacha using a teleporter location and deposit all traps.

    Parameters:
    - teleporter: position to teleport near/onto the gacha box
    - gacha_view_direction: the yaw/pitch or vector to face the gacha for inventory interactions
    """
    def __init__(self, bot_state: BotState, player_input: PlayerInput, inventory_manager: InventoryManager, teleporter, gacha_view_direction):
        self.bot_state: BotState = bot_state
        self.player_input: PlayerInput = player_input
        self.inventory_manager: InventoryManager = inventory_manager
        self.teleporter = teleporter
        self.gacha_view_direction = gacha_view_direction

    def run(self):
        self.player_input.log_task(self.bot_state, f"FeedGachaTask teleporter={self.teleporter}", task_obj=self)
        # Teleport to gacha vicinity, face gacha, then deposit
        self.player_input.teleport_to(self.teleporter, self.bot_state)
        self.player_input.look_at(self.gacha_view_direction, self.bot_state)
        time.sleep(0.5)  # wait for teleport stabilization
        self.inventory_manager.open_inv()
        self.inventory_manager.store_all()
        self.inventory_manager.close_inv()


class CollectCrystalsTask(BaseTask):
    """Collect crystals from the Pego at a Gacha Box.

    Parameters:
    - teleporter: location name to teleport near the gacha box
    - pego_view_direction: yaw/pitch to face the Pego for inventory interaction
    - count: optional number of collection iterations (default 1)
    """
    def __init__(self, bot_state: BotState, player_input: PlayerInput, inventory_manager: InventoryManager, teleporter, pego_view_direction, count: int = 1):
        self.bot_state = bot_state
        self.player_input = player_input
        self.inventory_manager = inventory_manager
        self.teleporter = teleporter
        self.pego_view_direction = pego_view_direction
        self.count = max(1, int(count))

    def run(self):
        self.player_input.log_task(self.bot_state, f"CollectCrystalsTask teleporter={self.teleporter} count={self.count}", task_obj=self)
        # Move to box and face the Pego
        self.player_input.teleport_to(self.teleporter, self.bot_state)
        self.player_input.look_at(self.pego_view_direction, self.bot_state)
        time.sleep(0.5)

        debug(f"Collecting Crystals from {self.teleporter}")
        self.inventory_manager.open_inv()
        self.inventory_manager.take_all()
        self.inventory_manager.close_inv()


class CrackCrystalsTask(BaseTask):
    """Crack crystals by teleporting to a grinder and processing crystals.

    Base routine:
    - Teleport to grinder location
    - Face grinder
    - Open grinder inventory and store crystals from player inventory
    - Repeat for `count`

    Parameters:
    - grinder_teleporter: location name for the grinder
    - grinder_view_direction: yaw/pitch to face the grinder
    - count: optional number of crack iterations (default 1)
    """
    def __init__(self, bot_state: BotState, player_input: PlayerInput, inventory_manager: InventoryManager, grinder_teleporter, grinder_view_direction, count: int = 1, resource_teleporter=None, resource_view_direction=None):
        self.bot_state = bot_state
        self.player_input = player_input
        self.inventory_manager = inventory_manager
        self.grinder_teleporter = grinder_teleporter
        self.grinder_view_direction = grinder_view_direction
        self.count = max(1, int(count))
        # Optional: where to deposit resources after cracking
        self.resource_teleporter = resource_teleporter
        self.resource_view_direction = resource_view_direction

    def run(self):
        self.player_input.log_task(self.bot_state, f"CrackCrystalsTask grinder={self.grinder_teleporter} count={self.count}", task_obj=self)
        # Move to grinder and face it
        self.player_input.teleport_to(self.grinder_teleporter, self.bot_state)
        #self.player_input.look_at(self.grinder_view_direction, self.bot_state)
        time.sleep(0.5)
        for i in range(self.count):
            debug(f"Cracking crystals at grinder iteration {i+1}/{self.count}")
            self.crack_crystals()
            time.sleep(1)
            self.sort_ressources()

    def crack_crystals(self):
        """Crack the crystals in the player's inventory by opening them.

        Flow:
        - Filter own inventory to crystals (e.g., 'gacha')
        - Click first slot of own inventory
                - Perform smooth left-right mouse movements while pressing 'e' to open crystals
                    in fixed cycles, then re-check if crystals remain; repeat until none left
        """
        debug("Opening Crystals...")
        inp = self.inventory_manager.input
        # Ensure own inventory is open
        self.inventory_manager.open_own_inv()
        # Filter left text field to 'gacha' to find crystals quicker
        if getattr(inp, 'textfield_left', None):
            inp.enter_text(inp.textfield_left, 'gacha')
            time.sleep(0.2)
        # Click first slot of own inventory if known
        if getattr(inp, 'first_slot_own', None):
            inp.move_mouse_absolute(*inp.first_slot_own)
            pyautogui.click()
        # Helper to check if crystals remain in first slot region
        def _crystal_left() -> bool:
            return bool(self.player_input.crystal_left(getattr(inp, 'first_slot_own_scan', None), confidence=0.4))
        # Movement parameters
        CYCLES = 2           # number of left-right cycles before re-checking
        SMOOTH_PIX = 540       # pixels per smooth move
        # Repeat cycles, then re-check if crystals remain; stop when none left
        while True:
            for _ in range(CYCLES):
                inp.smooth_mouse_move(SMOOTH_PIX, 0, key='e')
                inp.smooth_mouse_move(-SMOOTH_PIX, 0, key='e')
            if getattr(inp, 'nudge_mouse', None):
                inp.move_mouse_absolute(*inp.nudge_mouse)
                pyautogui.click()
            empty = not _crystal_left()
            inp.move_mouse_absolute(*inp.first_slot_own)
            if empty:
                break
        # Close own inventory before sorting phase
        self.inventory_manager.close_inv()
        debug("All Crystals cracked...sorting.")

    def sort_ressources(self):
        """Sort resources obtained from cracked crystals to configured dedis.

        Loads config from config/didi.json and deposits only into storages
        where is_gacha_crystal_ressource == true. For each such storage:
        - optionally teleport (if teleporter provided)
        - face the storage view_direction
        - open storage inventory and filter player inventory by resource name
        - store filtered items
        """
        # Load dedis configuration via PlayerInput helper
        data = self.player_input.load_json("didi.json")
        dedis = data.get("dedis", [])
        time.sleep(0.2)
        debug("Sorting cracked resources to dedis...")
        # Iterate configured storages flagged for gacha resources
        for d in dedis:
            try:
                if not d.get("is_gacha_crystal_ressource", False):
                    continue
                resource = d.get("resource")
                if not resource:
                    continue
                #tp = d.get("teleporter")
                view = d.get("view_direction") or d.get("view")
                # Optional: crouch requirement per storage
                required_crouch = bool(d.get("crouch", False))
                if required_crouch != bool(self.bot_state.is_crouching):
                    self.player_input.crouch(self.bot_state)
                #if tp:
                #    self.player_input.teleport_to(tp, self.bot_state)
                if view is not None:
                    self.player_input.look_at(view, self.bot_state)
                time.sleep(0.2)
                # Open storage inventory in front
                self.inventory_manager.open_inv()
                self.inventory_manager.store_all()
                self.inventory_manager.close_inv()
            except RestartTask as rt:
                # Propagate restart so this stage restarts from the beginning on next run
                warn(f"sort_ressources: restart while sorting resource '{d.get('resource')}': {rt}")
                raise
            except Exception as e:
                # Any error during sorting should restart the entire sort stage to re-cycle all storages
                warn(f"sort_ressources: error while sorting resource '{d.get('resource')}': {e}")
                raise RestartTask(f"sort_ressources failure at resource '{d.get('resource')}'")
        # Ensure we leave crouch state off after finishing all dedis to avoid impacting next tasks
        try:
            if bool(self.bot_state.is_crouching):
                self.player_input.crouch(self.bot_state)
        except Exception:
            pass

class SortLootAndGrindTask(BaseTask):
    """Sort loot (armor/weapons/etc.), grind junk, then sort resources from grinding.

    Parameters:
    - grinder_teleporter: teleporter near grinder
    - grinder_view_direction: yaw/pitch to face grinder
    - loot_storage_teleporter/view: optional destination to keep good loot
    - resource_storage_teleporter/view: optional destination for resources after grinding
    - keep_filters: optional list of strings to filter 'kept' loot (e.g., ["flak", "rifle"]). Base implementation uses simple text filters.
    """
    def __init__(self, bot_state: BotState, player_input: PlayerInput, inventory_manager: InventoryManager,
                 grinder_teleporter=None, grinder_view_direction=None,
                 loot_storage_teleporter=None, loot_storage_view_direction=None,
                 resource_storage_teleporter=None, resource_storage_view_direction=None,
                 keep_filters=None):
        self.bot_state = bot_state
        self.player_input = player_input
        self.inventory_manager = inventory_manager
        # Auto-resolve grinder teleporter/view if not provided
        self.grinder_teleporter = grinder_teleporter or (
            'grinder' if (hasattr(self.player_input, 'teleporters') and 'grinder' in self.player_input.teleporters)
            else getattr(self.player_input, 'teleporter_render', 'grinder')
        )
        self.grinder_view_direction = grinder_view_direction or (
            self.player_input.get_calibration_view_direction('grinder')
            or self.player_input.get_calibration_view_direction('grinder_view')
        )
        self.loot_storage_teleporter = loot_storage_teleporter
        self.loot_storage_view_direction = loot_storage_view_direction
        self.resource_storage_teleporter = resource_storage_teleporter
        self.resource_storage_view_direction = resource_storage_view_direction
        # Load keep filters from config if not provided
        if keep_filters and len(keep_filters) > 0:
            self.keep_filters = keep_filters
        else:
            data = self.player_input.load_json("loot_filters.json")
            self.keep_filters = data.get("keep_filters", [])
        # Auto-resolve loot storage teleporter/view if not provided
        if self.loot_storage_teleporter is None:
            # Default behavior: loot storage reachable at grinder TP
            self.loot_storage_teleporter = self.grinder_teleporter
        if self.loot_storage_view_direction is None:
            self.loot_storage_view_direction = (
                self.player_input.get_calibration_view_direction("loot_storage")
                or self.player_input.get_calibration_view_direction("poly_vault")
            )
        # Auto-resolve resource storage view if not provided (optional)
        if self.resource_storage_view_direction is None:
            self.resource_storage_view_direction = self.player_input.get_calibration_view_direction("resource_storage")

    def run(self):
        self.player_input.log_task(self.bot_state, "SortLootAndGrindTask", task_obj=self)
        # Go to grinder
        #self.player_input.teleport_to(self.grinder_teleporter, self.bot_state)
        #self.player_input.look_at(self.grinder_view_direction, self.bot_state)
        #time.sleep(0.3)
        # Sort loot: move 'kept' items to loot storage (if configured)
        self.sort_loot()
        # Grind remaining junk (base: deposit all to grinder to be processed)
        self.grind_junk()


    def sort_loot(self):
        """Keep desired loot and move it to a loot storage (if configured).

        Base approach: for each keep filter, set player's inventory filter then Store All into loot storage.
        """
        if not self.keep_filters:
            return
        if self.loot_storage_teleporter is None:
            warn("SortLootAndGrindTask: loot storage not configured; skipping loot sorting.")
            return
        if self.loot_storage_view_direction is not None:
            self.player_input.look_at(self.loot_storage_view_direction, self.bot_state)
        time.sleep(0.2)
        debug("Sorting poly loot to loot storage...")
        for flt in self.keep_filters:
            # Filter and store kept items
            self.inventory_manager.open_inv()
            try:
                inp = self.inventory_manager.input
                if getattr(inp, 'textfield_left', None):
                    inp.enter_text(inp.textfield_left, flt)
                self.inventory_manager.store_all()
                if getattr(inp, 'textfield_left', None):
                    inp.enter_text(inp.textfield_left, "")
            finally:
                if self.is_vault_full():
                    self.do_vault_full_task()
        self.inventory_manager.close_inv()
        time.sleep(0.5)

    def grind_junk(self):
        """Deposit remaining items to grinder for processing (base routine)."""
        # Ensure we are facing and opening the grinder explicitly
        self.player_input.look_at(self.grinder_view_direction, self.bot_state)
        time.sleep(0.2)
        debug("Opening grinder inventory for junk deposit...")
        self.inventory_manager.open_inv('grinder')
        try:
            self.inventory_manager.store_all()
            # After depositing, check slot count and trigger grind if over threshold
            try:
                time.sleep(0.5)  # wait for inventory to update after storing
                # Use configured grinder slot scan region
                region = getattr(self.player_input, 'grinder_slots', None)
                full = self.inventory_manager.grinder_slots(region, 80)
                if full:
                    self.click_grind_button()
                    self.inventory_manager.close_inv()
                    time.sleep(1.0)  # wait for grinding to complete
                    # Metal-first presorting loop, then general resource distribution
                    self.grind_inventory_metal_first()
                    self.sort_resources_from_grinding()
                else:
                    debug("Grinder slots <= 80; skipping grinding.")
                    self.inventory_manager.close_inv()
            except RestartTask as rt:
                warn(f"SortLootAndGrindTask: restart triggered during grind check: {rt}")
                raise
            except Exception as e:
                # Convert unexpected errors to restart so the sort stage is repeated
                warn(f"SortLootAndGrindTask: slot count/grind check failed: {e}")
                raise RestartTask(f"grind_junk check failure: {e}")
        except RestartTask:
            # Bubble up explicit restarts from called routines
            raise
        except Exception as e:
            # Any failure while depositing/opening should restart this stage
            warn(f"SortLootAndGrindTask: grinder deposit failed: {e}")
            try:
                self.inventory_manager.close_inv()
            except Exception:
                pass
            raise RestartTask(f"grind_junk deposit failure: {e}")
        finally:
            info("Grinded junk and sorted")

    def is_vault_full(self) -> bool:
        """Dummy check for whether the poly vault (loot storage) is full.

        Future implementation could read the slot count from UI via OCR.
        For now, always returns False.
        """
        # Placeholder: always not full
        return False

    def do_vault_full_task(self) -> None:
        """Dummy task executed when the vault is detected as full.

        Replace with a real workflow (e.g., switch to a secondary vault,
        alert the user, or move items elsewhere).
        """
        print("Vault appears full. Running dummy follow-up task...")
        time.sleep(0.5)
        os._exit(0)

    def click_grind_button(self) -> None:
        """Dummy: simulate clicking the grinder 'Grind' button.

        Replace with actual UI coordinates and input actions. Uses
        `config/click_positions.json` key `grind_button` if present.
        """
        debug("Clicking Grind button")
        try:
            pos = getattr(self.player_input, "grind_button", None)
            if pos is not None:
                x, y = pos
                self.player_input.move_mouse_absolute(x, y)
                pyautogui.click()
        except Exception as e:
            warn(f"Click Grind button failed: {e}")

    def sort_resources_from_grinding(self):
        """Move produced resources from grinder to configured dedis from didi.json.

        Loads config from `config/didi.json` and iterates storages flagged as
        grinder resource destinations (key `is_grinder_resource: true`). For each:
        - optionally teleport
        - face the storage view_direction
        - open inventory, store all
        """
        info("Sorting grinder resources...")
        time.sleep(0.4)
        # First, pull resources from grinder -> player
        self.player_input.look_at(self.grinder_view_direction, self.bot_state)
        self.inventory_manager.open_inv("grinder")
        self.inventory_manager.take_all()
        self.inventory_manager.close_inv()

        # Load dedis configuration
        data = self.player_input.load_json("didi.json")
        dedis = data.get("dedis", [])
        time.sleep(0.5)
        grinder_statistics = {}
        # Accumulate per-dedi stats to save once after loop
        stats_dir = os.path.join(os.path.dirname(__file__), "..", "logs")
        ts_now = time.strftime("%Y-%m-%d %H:%M:%S")
        # Build stable resource column order from config
        resource_names = []
        seen = set()
        for d in dedis:
            # Only include dedicated storages
            if not d.get("is_dedi", False):
                continue
            name = d.get("resource")
            if isinstance(name, str):
                up = name.upper()
                if up not in seen:
                    seen.add(up)
                    resource_names.append(up)
        # Map resource -> amount captured this run
        per_dedi_amounts = {res: 0 for res in resource_names}

        # Iterate over configured grinder resource storages
        for i, d in enumerate(dedis):
            try:
                
                view = d.get("view_direction")
                name = d.get("resource") 
                # Every 3 dedis, calibrate view
                if i > 0 and i % 3 == 0:
                    debug(f"Calibrating view after {i} dedis...")
                    self.player_input.calibrate_current_view(bot_state=self.bot_state)
                # Optional: crouch requirement per storage when distributing grinder outputs
                required_crouch = bool(d.get("crouch", False))
                if required_crouch != bool(self.bot_state.is_crouching):
                    self.player_input.crouch(self.bot_state)
                if view is not None:
                    self.player_input.look_at(view, self.bot_state)
                time.sleep(0.1)
                self.inventory_manager.open_inv(name if isinstance(name, str) else None)
                # Read dedicated storage amount banner if configured
                region = getattr(self.player_input, 'dedi_amt_scan', None)
                # Only capture amounts for true dedicated storages
                if region and d.get("is_dedi", False):
                    amt, _res_ignored = self.player_input.ocr.read_dedi_amount(region)
                    if amt is not None:
                        res_name = (name or '').upper() if isinstance(name, str) else ''
                        debug(f"Dedi '{res_name}': CONTAINS {amt}")
                        if res_name:
                            per_dedi_amounts[res_name] = amt
                self.inventory_manager.store_all()
                self.inventory_manager.close_inv()
                time.sleep(0.2)
            except Exception as e:
                warn(f"sort_resources_from_grinding: error at '{d.get('name') or d.get('resource')}': {e}")
        # Ensure we leave crouch state off at the end to not affect subsequent tasks
        try:
            if bool(self.bot_state.is_crouching):
                self.player_input.crouch(self.bot_state)
        except Exception:
            pass

        # Save accumulated stats once after sorting all dedis
        if any(per_dedi_amounts.values()):
            try:
                os.makedirs(stats_dir, exist_ok=True)
                # Prefer Excel if openpyxl is available; else fallback to CSV
                try:
                    import openpyxl
                    from openpyxl import Workbook
                    xlsx_path = os.path.join(stats_dir, "grind_stats.xlsx")
                    # Open or create workbook
                    if os.path.exists(xlsx_path):
                        wb = openpyxl.load_workbook(xlsx_path)
                        ws = wb.active
                        # If header doesn't match desired format, rebuild
                        if ws.max_row == 0 or ws.max_column == 0:
                            ws.append(["Timestamp"] + resource_names)
                    else:
                        wb = Workbook()
                        ws = wb.active
                        ws.append(["Timestamp"] + resource_names)  # header
                    # Ensure header has all resources; if new resources appear, extend header
                    header = [cell.value for cell in ws[1]]
                    desired_header = ["Timestamp"] + resource_names
                    if header != desired_header:
                        # Rebuild header to match desired order
                        ws.delete_rows(1, 1)
                        ws.insert_rows(1)
                        ws.append(desired_header)
                    # Build row for current run
                    row = [ts_now] + [per_dedi_amounts.get(res, 0) for res in resource_names]
                    ws.append(row)
                    wb.save(xlsx_path)
                    debug(f"Saved grind stats to Excel: {xlsx_path}")
                except Exception as e:
                    # Fallback to CSV
                    csv_path = os.path.join(stats_dir, "grind_stats.csv")
                    new_file = not os.path.exists(csv_path)
                    with open(csv_path, "a", encoding="utf-8") as f:
                        if new_file:
                            f.write(",".join(["Timestamp"] + resource_names) + "\n")
                        row = [ts_now] + [str(per_dedi_amounts.get(res, 0)) for res in resource_names]
                        f.write(",".join(row) + "\n")
                    info(f"Saved grind stats to CSV: {csv_path} (Excel unavailable: {e})")
            except Exception as e:
                warn(f"Failed to save accumulated grind stats: {e}")

    def store_metal(self):
        """Store metal, wood, and stone to their dedicated storage(s) defined in didi.json.

        Expects dedis entries with resource == 'metal', 'wood', or 'stone' and optional view_direction.
        """
        data = self.player_input.load_json("didi.json")
        dedis = data.get("dedis", [])
        for d in dedis:
            try:
                resource = (d.get("resource") or '').lower()
                if resource not in ('metal', 'wood'):
                    continue
                view = d.get("view_direction")
                if view is not None:
                    self.player_input.look_at(view, self.bot_state)
                time.sleep(0.1)
                self.inventory_manager.open_inv(resource)
                self.inventory_manager.store_all()
                self.inventory_manager.close_inv()
                time.sleep(1)
            except Exception as e:
                warn(f"store_metal: error at {resource} dedi '{d.get('name') or d.get('resource')}': {e}")
        time.sleep(1)

    def grind_inventory_metal_first(self):
        """After grinding: repeatedly take grinder contents and store only metal, wood and stone until first slot empty."""
        debug("Metal-first presorting loop starting...")
        # Align and pull initial batch
        self.player_input.look_at(self.grinder_view_direction, self.bot_state)
        self.inventory_manager.open_inv('grinder')
        self.inventory_manager.take_all()
        tries = 0
        max_tries = 8
        while True:
            empty = self.player_input.slot_empty(getattr(self.player_input, 'first_slot_grinder_scan', None))
            if empty:
                info("Grinder first slot empty – presort complete.")
                self.inventory_manager.close_inv()
                time.sleep(1)
                break
            tries += 1
            debug(f"Grinder still has items; storing metal chunk... (try {tries}/{max_tries})")
            # Close grinder, store metal, then re-open and take next batch
            self.inventory_manager.close_inv()
            time.sleep(2)
            self.player_input.get_calibration_view_direction('grinder')
            self.store_metal()
            # Failsafe: after max_tries, run resource sorting once to relieve pressure
            if tries == max_tries:
                try:
                    warn("Failsafe triggered: running sort_resources_from_grinding() once.")
                    self.sort_resources_from_grinding()
                except Exception as e:
                    warn(f"Failsafe sort_resources_from_grinding failed: {e}")
                # If still not empty after failsafe, restart sort stage to recover
                self.player_input.look_at(self.grinder_view_direction, self.bot_state)
                self.inventory_manager.open_inv('grinder')
                still_full = not self.player_input.slot_empty(getattr(self.player_input, 'first_slot_grinder_scan', None))
                if still_full:
                    self.inventory_manager.close_inv()
                    time.sleep(1)
                    raise RestartTask("grind_inventory_metal_first stuck after failsafe")
            # Re-open grinder for next batch
            self.player_input.look_at(self.grinder_view_direction, self.bot_state)
            self.inventory_manager.open_inv('grinder')
            self.inventory_manager.take_all()
            time.sleep(0.15)
        debug("Metal-first presorting loop finished.")


class CollectAndCrackAllGachasTask(BaseTask):
    """High-level: for each gacha box, collect Pego crystals and crack them at a grinder.

    Expects `gacha_boxes` to be a list of dicts with:
      {
        'teleporter': <box teleporter>,
        'pego_view_direction': <yaw,pitch>,
      }
    And grinder info:
      grinder_teleporter: <teleporter name>,
      grinder_view_direction: <yaw,pitch>
    """
    def __init__(self, bot_state: BotState, player_input: PlayerInput, inventory_manager: InventoryManager, gacha_boxes=None, grinder_teleporter=None, grinder_view_direction=None, per_box_collect_count: int = 1, per_box_crack_count: int = 1):
        self.bot_state = bot_state
        self.player_input = player_input
        self.inventory_manager = inventory_manager
        # Load all boxes dynamically when not provided
        self.gacha_boxes = gacha_boxes if gacha_boxes is not None else load_gacha_boxes_from_config()
        self.grinder_teleporter = grinder_teleporter
        self.grinder_view_direction = grinder_view_direction
        self.per_box_collect_count = max(1, int(per_box_collect_count))
        self.per_box_crack_count = max(1, int(per_box_crack_count))

    def run(self):
        # If boxes list is empty, attempt to load from config again
        if not self.gacha_boxes:
            self.gacha_boxes = load_gacha_boxes_from_config()
        # Guard: no boxes
        if not self.gacha_boxes:
            warn("CollectAndCrackAllGachasTask: no gacha boxes found in config.")
            return

        # Resume support via bot_state
        start_idx = int(getattr(self.bot_state, 'collect_checkpoint_idx', 0) or 0)
        # Do NOT capture start_stage once; re-read from bot_state each box so
        # once a resumed stage completes and is cleared, subsequent boxes run fully.
        stages = ('collect', 'crack', 'sort')

        for idx, box in enumerate(self.gacha_boxes):
            if idx < start_idx:
                continue
            teleporter = box.get('teleporter')
            pego_view = box.get('pego_view_direction')
            if not teleporter or pego_view is None:
                continue
            # Re-evaluate current stage from bot_state for this iteration
            start_stage = getattr(self.bot_state, 'collect_checkpoint_stage', None)

            def run_stage(stage_name, func):
                # set checkpoint and execute stage, re-raise RestartTask to abort
                self.bot_state.collect_checkpoint_idx = idx
                self.bot_state.collect_checkpoint_stage = stage_name
                debug(f"Checkpoint set: idx={idx+1}/{len(self.gacha_boxes)} stage='{stage_name}'")
                try:
                    func()
                    # Clear stage after successful completion of this stage
                    self.bot_state.collect_checkpoint_stage = None
                    debug(f"Checkpoint stage cleared after '{stage_name}' for box {idx+1}")
                except RestartTask as rt:
                    warn(f"CollectAndCrackAllGachasTask: restart at box {idx+1} stage '{stage_name}': {rt}")
                    raise
                except Exception as e:
                    # Convert unexpected errors into RestartTask so the driver can retry cleanly
                    warn(f"CollectAndCrackAllGachasTask: unexpected error at box {idx+1} stage '{stage_name}': {e}")
                    raise RestartTask(f"unexpected error during {stage_name}: {e}")

            # Collect Pego crystals
            if start_stage is None or start_stage == stages[0]:
                info(f"Collecting box {idx+1}/{len(self.gacha_boxes)}")
                run_stage('collect', lambda: CollectCrystalsTask(
                    self.bot_state,
                    self.player_input,
                    self.inventory_manager,
                    teleporter,
                    pego_view,
                    count=self.per_box_collect_count
                ).run())

            # Crack crystals at grinder
            if start_stage is None or start_stage in (stages[1], stages[0]):
                info(f"Cracking box {idx+1}/{len(self.gacha_boxes)}")
                run_stage('crack', lambda: CrackCrystalsTask(
                    self.bot_state,
                    self.player_input,
                    self.inventory_manager,
                    self.grinder_teleporter,
                    self.grinder_view_direction,
                    count=self.per_box_crack_count
                ).run())

            # Sort loot and grind resources produced
            info(f"Sorting box {idx+1}/{len(self.gacha_boxes)}")
            run_stage('sort', lambda: SortLootAndGrindTask(
                self.bot_state,
                self.player_input,
                self.inventory_manager
            ).run())

            # Completed box; advance checkpoint to next and clear stage
            self.bot_state.collect_checkpoint_idx = idx + 1
            self.bot_state.collect_checkpoint_stage = None
            debug(f"Checkpoint advanced to next box: idx={self.bot_state.collect_checkpoint_idx}")


class FertilizePlotsTask(BaseTask):
    def __init__(self, bot_state: BotState, player_input: PlayerInput, inventory_manager: InventoryManager, crop_plots):
        self.bot_state: BotState = bot_state
        self.player_input: PlayerInput = player_input
        self.inventory_manager: InventoryManager = inventory_manager
        self.crop_plots = crop_plots

    def run(self):
        self.player_input.log_task(self.bot_state, "FertilizePlotsTask start", task_obj=self)
        for plot in self.crop_plots:
            self.player_input.look_at(plot['view_direction'], self.bot_state)
            self.inventory_manager.open_inv()
            self.inventory_manager.store_all()
            self.inventory_manager.close_inv()


class MajorTask:
    """Sequences multiple base tasks for complex workflows."""
    def __init__(self, tasks):
        self.tasks = tasks

    def run(self):
        for task in self.tasks:
            task.run()
        # (intentionally left blank)


class FeedAllGachasMajorTask(BaseTask):
    """High-level workflow: for each gacha set, gather traps then feed both gachas.

    Expects `gacha_sets` to be an iterable where each item is a dict:
      {
        'tp_box': <position for box teleporter>,
        'tp_plots_gacha1': <teleporter to plots for gacha1>,
        'tp_plots_gacha2': <teleporter to plots for gacha2>,
        'gacha1_view_direction': <dir>,
        'gacha2_view_direction': <dir>
      }
    """
    def __init__(self, bot_state: BotState, player_input: PlayerInput, inventory_manager: InventoryManager, gacha_sets=None):
        self.bot_state = bot_state
        self.player_input = player_input
        self.inventory_manager = inventory_manager
        self.gacha_sets = gacha_sets

    def run(self):
        import os, json
        # Load gacha_sets.json when not provided
        sets = self.gacha_sets
        if sets is None:
            cfg_path = os.path.join(os.path.dirname(__file__), "..", "config", "gacha_sets.json")
            with open(cfg_path, "r") as f:
                data = json.load(f)
            sets = data.get("gacha_sets", [])
        if not sets:
            warn("FeedAllGachasMajorTask: no gacha_sets found in config.")
            return

        # Load crop plot look positions
        crop_plot_path = os.path.join(os.path.dirname(__file__), "..", "config", "crop_plot_look_positions.json")
        with open(crop_plot_path, "r") as f:
            crop_plot_cfg = json.load(f)
        crop_plots = crop_plot_cfg.get("crop_plot_look_positions", [])
        if len(crop_plots) < 32:
            warn(f"Expected 32 crop plot look positions; found {len(crop_plots)}")
            return
        plots = [{
            "view_direction": item.get("view_direction"),
            "crouch": bool(item.get("crouch", False)),
        } for item in crop_plots[:32]]

        # Resume support: box index and stage
        start_idx = int(getattr(self.bot_state, 'major_checkpoint_idx', 0) or 0)
        start_stage = getattr(self.bot_state, 'major_checkpoint_stage', None)
        stages = ('gettraps', 'feed_gacha1', 'feed_gacha2')

        for idx, box in enumerate(sets):
            if idx < start_idx:
                continue
            # Always reset last_action_success at the start of each box
            self.bot_state.last_action_success = True
            # Cycle through 4 sets of plots: plots1, plots2, plots3, plots4
            plots_cycle = ["plots1", "plots2", "plots3", "plots4"]
            tp_plots = plots_cycle[idx % 4]
            tp_box = box.get("tp_box")
            g1_view = box.get("gacha1_view_direction") or box.get("gacha1", {}).get("view_direction")
            g2_view = box.get("gacha2_view_direction") or box.get("gacha2", {}).get("view_direction")

            def run_stage(stage_name, func):
                self.bot_state.major_checkpoint_idx = idx
                self.bot_state.major_checkpoint_stage = stage_name
                try:
                    func()
                except RestartTask as rt:
                    warn(f"FeedAllGachasMajorTask: restart at box {idx+1} stage '{stage_name}': {rt}")
                    # Keep checkpoint and abort major run to resume next time
                    raise

            info(f"Feeding Box {idx+1}/{len(sets)} using {tp_plots}")
            # 1. Get traps from selected plots
            if start_stage is None or start_stage == stages[0]:
                run_stage('gettraps', lambda: (
                    self.player_input.teleport_to(tp_plots, self.bot_state) if tp_plots else None,
                    GetTrapsTask(self.bot_state, self.player_input, self.inventory_manager, crop_plots=plots, indices=range(0, 32)).run()
                ))
            # Always run FeedGachaTask after GetTrapsTask, unless a RestartTask was raised
            # 2. Teleport to box and feed Gacha 1
            run_stage('feed_gacha1', lambda: (
                self.player_input.teleport_to(tp_box, self.bot_state) if tp_box else None,
                FeedGachaTask(self.bot_state, self.player_input, self.inventory_manager, tp_box, g1_view).run() if (tp_box and g1_view) else None
            ))
            # 3. Feed Gacha 2 (already at box)
            run_stage('feed_gacha2', lambda: (
                FeedGachaTask(self.bot_state, self.player_input, self.inventory_manager, tp_box, g2_view).run() if (tp_box and g2_view) else None
            ))

            # Completed box; advance checkpoint
            self.bot_state.major_checkpoint_idx = idx + 1
            self.bot_state.major_checkpoint_stage = None

        # Removed end-of-task drop_all cleanup to prevent unintended loops

