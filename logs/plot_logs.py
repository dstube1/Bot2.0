import os
import sys
from datetime import datetime


def load_data(logs_dir):
	"""Load grind stats from Excel (preferred) or CSV.

	Returns (timestamps: list[datetime], series: dict[str, list[float]], columns: list[str])
	where columns is the ordered list of resource names.
	"""
	xlsx_path = os.path.join(logs_dir, "grind_stats.xlsx")
	csv_path = os.path.join(logs_dir, "grind_stats.csv")

	# Prefer Excel if available
	if os.path.exists(xlsx_path):
		try:
			import openpyxl
			wb = openpyxl.load_workbook(xlsx_path, data_only=True)
			ws = wb.active
			# Expect header row: Timestamp, RES1, RES2, ...
			header = [c.value for c in ws[1]]
			if not header or header[0] != "Timestamp":
				raise ValueError("Unexpected header in grind_stats.xlsx")
			cols = header[1:]
			timestamps = []
			series = {col: [] for col in cols}
			for row in ws.iter_rows(min_row=2, values_only=True):
				if not row or all(v is None for v in row):
					continue
				ts_str = row[0]
				if isinstance(ts_str, datetime):
					ts = ts_str
				else:
					ts = datetime.strptime(str(ts_str), "%Y-%m-%d %H:%M:%S")
				timestamps.append(ts)
				for idx, col in enumerate(cols, start=1):
					val = row[idx] if idx < len(row) else 0
					try:
						series[col].append(float(val or 0))
					except Exception:
						series[col].append(0.0)
			return timestamps, series, cols
		except Exception as e:
			print(f"Failed reading Excel ({xlsx_path}), will try CSV: {e}")

	# CSV fallback
	if os.path.exists(csv_path):
		import csv
		timestamps = []
		series = {}
		cols = []
		with open(csv_path, "r", encoding="utf-8") as f:
			reader = csv.reader(f)
			header = next(reader, None)
			if not header or header[0] != "Timestamp":
				raise ValueError("Unexpected header in grind_stats.csv")
			cols = header[1:]
			series = {col: [] for col in cols}
			for row in reader:
				if not row:
					continue
				ts = datetime.strptime(row[0], "%Y-%m-%d %H:%M:%S")
				timestamps.append(ts)
				for idx, col in enumerate(cols, start=1):
					try:
						series[col].append(float(row[idx]))
					except Exception:
						series[col].append(0.0)
		return timestamps, series, cols

	raise FileNotFoundError("No grind_stats.xlsx or grind_stats.csv found in logs directory")


def _smooth_random_zeros(values):
	"""Replace isolated zeros with the average of their neighbors.

	Only applies when both previous and next values exist. Does not modify
	leading/trailing zeros or sequences of zeros.
	"""
	if not values:
		return values
	smoothed = list(values)
	for i in range(1, len(values) - 1):
		if smoothed[i] == 0 and smoothed[i - 1] != 0 and smoothed[i + 1] != 0:
			smoothed[i] = (smoothed[i - 1] + smoothed[i + 1]) / 2.0
	return smoothed


def plot_series(timestamps, series, cols, output_path=None):
	# Lazy import matplotlib to avoid dependency if just loading
	try:
		import matplotlib.pyplot as plt
		import matplotlib.dates as mdates
	except Exception as e:
		print("matplotlib is required to plot. Install with: pip install matplotlib")
		raise

	# Separate dust series from others (case-insensitive)
	dust_keys = {"dust1", "dust2", "dust3"}
	dust_cols = [c for c in cols if c.lower() in dust_keys]
	other_cols = [c for c in cols if c.lower() not in dust_keys]

	# Build two vertically stacked subplots sharing the X axis
	fig, (ax_top, ax_bottom) = plt.subplots(2, 1, sharex=True, figsize=(12, 8))

	# Top: dust series
	if dust_cols:
		for col in dust_cols:
			vals = _smooth_random_zeros(series.get(col, []))
			ax_top.step(timestamps, vals, where='post', label=col)
		ax_top.set_ylabel("Amount (dust1/2/3)")
		ax_top.legend(loc="best", ncol=2)
		ax_top.grid(True, alpha=0.3)
	else:
		ax_top.set_visible(False)

	# Bottom: other resources
	if other_cols:
		for col in other_cols:
			vals = _smooth_random_zeros(series.get(col, []))
			ax_bottom.step(timestamps, vals, where='post', label=col)
		ax_bottom.set_ylabel("Amount (others)")
		ax_bottom.legend(loc="best", ncol=2)
		ax_bottom.grid(True, alpha=0.3)
	else:
		ax_bottom.set_visible(False)

	# Common X axis formatting
	ax_bottom.set_xlabel("Time")
	ax_bottom.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d\n%H:%M"))

	fig.suptitle("Grind Resources Over Time")
	fig.tight_layout(rect=[0, 0.03, 1, 0.97])

	if output_path:
		fig.savefig(output_path, dpi=150)
		print(f"Saved plot to: {output_path}")

	# Also show interactive window
	plt.show()


def main():
	# Allow running from anywhere
	script_dir = os.path.dirname(os.path.abspath(__file__))
	logs_dir = script_dir

	try:
		timestamps, series, cols = load_data(logs_dir)
	except Exception as e:
		print(f"Error loading data: {e}")
		sys.exit(1)

	output_path = os.path.join(logs_dir, "grind_stats.png")
	try:
		plot_series(timestamps, series, cols, output_path=output_path)
	except Exception as e:
		print(f"Error plotting data: {e}")
		sys.exit(1)


if __name__ == "__main__":
	main()

